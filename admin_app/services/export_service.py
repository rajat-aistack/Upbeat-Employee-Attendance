"""
Export service for generating Excel, CSV, and PDF reports.
"""
import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from admin_app.config import EXPORTS_DIR

logger = logging.getLogger(__name__)


def export_to_csv(data: List[Dict[str, Any]], filename: str = None) -> str:
    """Export attendance data to CSV file. Returns file path."""
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendance_report_{timestamp}.csv"

    filepath = EXPORTS_DIR / filename

    headers = [
        "Employee Name", "Date", "Check In", "Check Out",
        "Working Hours", "Status", "Late", "Missing Checkout", "Remarks"
    ]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for record in data:
            check_in = _format_datetime(record.get("check_in_at"))
            check_out = _format_datetime(record.get("check_out_at"))

            writer.writerow([
                record.get("employee_name", ""),
                record.get("date", ""),
                check_in,
                check_out,
                record.get("working_hours", ""),
                record.get("status", ""),
                "Yes" if record.get("is_late") else "No",
                "Yes" if record.get("missing_checkout") else "No",
                record.get("remarks", ""),
            ])

    logger.info(f"CSV exported to {filepath}")
    return str(filepath)


def export_to_excel(data: List[Dict[str, Any]], filename: str = None) -> str:
    """Export attendance data to Excel file. Returns file path."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.warning("openpyxl not installed, falling back to CSV")
        return export_to_csv(data, filename.replace(".xlsx", ".csv") if filename else None)

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendance_report_{timestamp}.xlsx"

    filepath = EXPORTS_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Style definitions
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Title row
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Upbeat Exposition Company — Attendance Report"
    title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="4F46E5")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    date_cell = ws["A2"]
    date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    date_cell.font = Font(name="Segoe UI", size=10, color="6B7280")
    date_cell.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Employee Name", "Date", "Check In", "Check Out",
        "Working Hours", "Status", "Late", "Missing Checkout", "Remarks"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    late_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    missing_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for row_idx, record in enumerate(data, 5):
        check_in = _format_datetime(record.get("check_in_at"))
        check_out = _format_datetime(record.get("check_out_at"))
        is_late = record.get("is_late", False)
        missing = record.get("missing_checkout", False)

        values = [
            record.get("employee_name", ""),
            record.get("date", ""),
            check_in,
            check_out,
            record.get("working_hours", ""),
            record.get("status", ""),
            "Yes" if is_late else "No",
            "Yes" if missing else "No",
            record.get("remarks", ""),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

            if is_late:
                cell.fill = late_fill
            elif missing:
                cell.fill = missing_fill

    # Auto-width columns
    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(4, ws.max_row + 1)
        )
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 30)

    wb.save(filepath)
    logger.info(f"Excel exported to {filepath}")
    return str(filepath)


def export_to_pdf(data: List[Dict[str, Any]], filename: str = None) -> str:
    """Export attendance data to PDF. Falls back to CSV if reportlab unavailable."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        logger.warning("reportlab not installed, falling back to CSV")
        return export_to_csv(data, filename.replace(".pdf", ".csv") if filename else None)

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"attendance_report_{timestamp}.pdf"

    filepath = EXPORTS_DIR / filename
    doc = SimpleDocTemplate(str(filepath), pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#4F46E5"),
    )
    elements.append(Paragraph("Upbeat Exposition Company", title_style))
    elements.append(Paragraph("Attendance Report", styles["Heading2"]))
    elements.append(Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.3 * inch))

    # Table
    table_data = [["Employee", "Date", "Check In", "Check Out", "Hours", "Status", "Late"]]

    for record in data:
        check_in = _format_datetime(record.get("check_in_at"))
        check_out = _format_datetime(record.get("check_out_at"))
        table_data.append([
            record.get("employee_name", ""),
            record.get("date", ""),
            check_in,
            check_out,
            record.get("working_hours", ""),
            record.get("status", ""),
            "Yes" if record.get("is_late") else "No",
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))
    elements.append(table)

    doc.build(elements)
    logger.info(f"PDF exported to {filepath}")
    return str(filepath)


def _format_datetime(dt_str) -> str:
    """Format an ISO datetime string for display."""
    if not dt_str:
        return ""
    try:
        if isinstance(dt_str, str):
            dt = datetime.fromisoformat(dt_str)
        else:
            dt = dt_str
        return dt.strftime("%I:%M %p")
    except Exception:
        return str(dt_str)
